# Importações
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import tkinter as tk
from tkinter import simpledialog
from tkinter.filedialog import askopenfile
from tkinter.filedialog import askdirectory
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinterdnd2 import DND_FILES, TkinterDnD
import openpyxl as xl
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
import sys
import os
import pandas as pd
import threading  # Importar threading
from tkinter import ttk  # Importar ttk para a barra de progresso

def main():
    try:
        caminho_pasta = var_caminho_pasta.get() 
        df_apuracao = encontrar_caminho_apuracao()
        df_orgao_sigla = encontrar_caminho_orgao_sigla()
        df_prateleira = encontrar_caminho_gabarito()
    
        if df_apuracao is None or df_prateleira is None or df_orgao_sigla is None:
            messagebox.showerror("Erro", "Um ou mais arquivos corretos não foram encontrados na pasta.")
            raise Exception('Um dos arquivos não foram encontrados')  # Isso vai parar o código

        # **Mostrar o indicador de carregamento**
        show_loading_indicator()

        # Retirando os possíveis espaços extras
        df_orgao_sigla['Cliente Nome'] = df_orgao_sigla['Cliente Nome'].apply(lambda x: x.strip() if isinstance(x, str) else x)
        df_prateleira['Código do item AVMG'] = df_prateleira['Código do item AVMG'].apply(lambda x: x.strip() if isinstance(x, str) else x)
        df_apuracao['Órgão/Entidade'] = df_apuracao['Órgão/Entidade'].apply(lambda x: x.strip() if isinstance(x, str) else x)

        # Etapa de preenchimento da coluna ajustes
        df_apuracao['Ajustes'] = '0'

        # Etapa de preenchimento de uma nova coluna SIGLA com valores da tabela gabarito órgão sigla
      
        df_orgao_sigla = df_orgao_sigla.rename(columns={'Cliente Nome': 'Órgão/Entidade'})
        df_apuracao = df_apuracao.merge(df_orgao_sigla[['Órgão/Entidade', 'Sigla']], on='Órgão/Entidade', how='left')
        df_apuracao = df_apuracao.drop_duplicates()

        # Etapa de preenchimento de uma nova coluna Item e Categoria com valores da tabela gabarito prateleira
        df_prateleira.drop(['Código Item Material - Numérico', 'Item Material', 'Item Correspondente', 'Situação'], axis='columns', inplace=True)
        df_prateleira.drop_duplicates(subset=['Código do item AVMG'], inplace=True)
        df_prateleira = df_prateleira.rename(columns={'Código do item AVMG': 'ID Item'})
        df_prateleira = df_prateleira.rename(columns={'Item Correspondente.1': 'Item'})
        df_apuracao = df_apuracao.merge(df_prateleira, on='ID Item', how='left')
        df_apuracao = df_apuracao.drop_duplicates()

        # Etapa preenchendo a coluna “Observação” com o texto “Sem observação”
        df_apuracao['Observação'] = 'Sem observação'

        # Etapa preenchendo a coluna “Exercício” com o ano referente ao da coluna Data da Aprovação
        df_apuracao['Exercício'] = df_apuracao['Data da Aprovação'].dt.year.fillna(0).astype(int)

        # Etapa de exclusão de valores "Não se aplica"
        df_apuracao['Data limite de entrega - pedido original'] = df_apuracao['Data limite de entrega - pedido original'].replace('Não se aplica', "")
        df_apuracao['Dias de atraso - pedido original'] = df_apuracao['Dias de atraso - pedido original'].replace('Não se aplica', "")
        df_apuracao['Data limite de entrega - entrega corretiva'] = df_apuracao['Data limite de entrega - entrega corretiva'].replace('Não se aplica', "")
        df_apuracao['Dias de atraso - entrega corretiva'] = df_apuracao['Dias de atraso - entrega corretiva'].replace('Não se aplica', "")

        # Verificar células em branco
        columns_to_check = ['Sigla', 'Categoria', 'Item']
        missing_columns = []

        for column in columns_to_check:
            if column in df_apuracao.columns:
                if df_apuracao[column].isnull().any():
                    missing_columns.append(column)
            else:
                print(f"Coluna {column} não encontrada no DataFrame df_apuracao.")

        if missing_columns:
            missing_message = f"\n\nObservação: as seguintes colunas possuem células em branco: {', '.join(missing_columns)}"
        else:
            missing_message = ""

        # Salvando o arquivo em Excel com algumas colunas em formato de data e design em geral
        caminho_arquivo_final = os.path.join(caminho_pasta, 'Valor Faturado.xlsx')
        with pd.ExcelWriter(caminho_arquivo_final, engine='openpyxl') as writer:
            df_apuracao.to_excel(writer, index=False, sheet_name='Valor Faturado')

        wb = load_workbook(caminho_arquivo_final)
        ws = wb['Valor Faturado']

        # Etapa de formatar datas
        colunas_de_data = ['Data do Fato Gerador'.strip(), 'Data da Aprovação'.strip(), 'Data do Ateste'.strip(), 'Data limite de entrega - entrega corretiva'.strip(), 'Data limite de entrega - pedido original'.strip()]
        for data in colunas_de_data:
            df_apuracao[data] = pd.to_datetime(df_apuracao[data]).dt.strftime('%d/%m/%Y')

        if 'date_style' not in wb.named_styles:
            date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
            wb.add_named_style(date_style)

        colunas_de_data_excel = ['BI', 'AV', 'AR', 'AQ']
        for col in colunas_de_data:
            if col in df_apuracao.columns:
                df_apuracao[col] = pd.to_datetime(df_apuracao[col], format='%d/%m/%Y', errors='coerce')

        for col in colunas_de_data_excel:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value:
                    cell.style = 'date_style'

        # Adicionando formatação aos nomes das colunas
        dark_blue_fill = PatternFill(start_color='00003366', end_color='00003366', fill_type='solid')  # Azul escuro
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Amarelo

        white_font_bold = Font(name='Arial', size=14, color='FFFFFF', bold=True)  # Texto branco em negrito
        black_font_bold = Font(name='Arial', size=14, color='000000', bold=True)  # Texto preto em negrito

        alignment = Alignment(horizontal='center', vertical='bottom')  # Centralizado e alinhado embaixo

        # Colunas que devem ser preenchidas em amarelo
        yellow_columns = {'AB', 'AI', 'AN'}
        # Atualizar para incluir todas as colunas de 'AQ' até a última coluna da planilha
        yellow_columns.update(
            [get_column_letter(col_num) for col_num in range(column_index_from_string('AQ'), ws.max_column + 1)]
        )

        # Definindo a altura da linha 1
        ws.row_dimensions[1].height = 100

        # Definindo a largura das colunas para 15 e aplicando formatação
        for col_num in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
            cell = ws[f'{col_letter}1']
            cell.alignment = alignment
            if col_letter in yellow_columns:
                cell.fill = yellow_fill
                cell.font = black_font_bold
            else:
                cell.fill = dark_blue_fill
                cell.font = white_font_bold

        # Aplicar filtro automático
        filtro = f"A1:{get_column_letter(ws.max_column)}1"
        ws.auto_filter.ref = filtro

        wb.save(caminho_arquivo_final)

        # **Ocultar o indicador de carregamento**
        hide_loading_indicator()

        # Exibir mensagem de sucesso
        messagebox.showinfo('Sucesso', f'Arquivo Valor Faturado salvo na pasta "{os.path.basename(caminho_pasta)}" com sucesso.{missing_message}') 

    except Exception as e:
        # **Ocultar o indicador de carregamento em caso de erro**
        hide_loading_indicator()

        messagebox.showerror("Erro", f"Erro ao gerar o arquivo. Verifique se os valores da planilha seguem o padrão recomendado.\n\nDetalhes do erro: {str(e)}")
        raise  # Relevanta a exceção para depuração
    finally:
        botao_processar.config(state='normal')

    return

def show_loading_indicator():
    # Exibir o indicador de carregamento
    loading_label['text'] = "Processando..."
    loading_label.grid(row=4, column=1, sticky='nsew')
    progress_bar.grid(row=5, column=1, padx=10, pady=5, sticky='nsew')
    progress_bar.start()
    janela.update_idletasks()

def hide_loading_indicator():
    # Ocultar o indicador de carregamento
    progress_bar.stop()
    loading_label['text'] = ""
    loading_label.grid_remove()
    progress_bar.grid_remove()
    janela.update_idletasks()

def iniciar_processamento():
    botao_processar.config(state='disabled')
    # Inicia a thread que executa a função main()
    threading.Thread(target=main).start()

def encontrar_caminho_apuracao():
    caminho_pasta = var_caminho_pasta.get()
    nome_apuracao = [f for f in os.listdir(caminho_pasta) if (f.startswith('Apuração')
                         or f.startswith('Apuração Faturamento')
                         or f.startswith('Apuracao do Faturamento')
                         or f.startswith('Apuracao faturamento')
                         or f.startswith('Apuração do faturamento'))
                         and f.endswith('xlsx')]

    if not nome_apuracao:
        print("Nenhum arquivo de 'Apuração do faturamento' encontrado.")
        return None  # Retorna None se não encontrar arquivos

    try:
        for arquivo in nome_apuracao:
            caminho_apuracao = os.path.join(caminho_pasta, arquivo)
            df_apuracao = pd.read_excel(caminho_apuracao, sheet_name='Apuração do Faturamento')
            df_apuracao.columns = df_apuracao.columns.str.strip()

            print(f"Arquivo {arquivo} lido com sucesso")
        return df_apuracao
    except PermissionError:
        messagebox.showerror("Erro", "Erro ao ler os arquivos. Verifique se o arquivo Apuração do Faturamento não esteja aberto")
        raise Exception('Erro ao ler arquivo Apuração do Faturamento')

def encontrar_caminho_gabarito():
    caminho_pasta = var_caminho_pasta.get()
    nome_gabarito = [f for f in os.listdir(caminho_pasta) if (f.startswith('Gabarito Prateleira')
                         or f.startswith('gabarito prateleira')
                         or f.startswith('Gabarito-Prateleira')
                         or f.startswith('gabarito-prateleira'))
                         and f.endswith('xlsx')]

    if not nome_gabarito:
        print("Nenhum arquivo de 'Gabarito Prateleira' encontrado.")
        return None

    try:
        for arquivo in nome_gabarito:
            caminho_gabarito = os.path.join(caminho_pasta, arquivo)
            df_prateleira = pd.read_excel(caminho_gabarito)
            df_prateleira.columns = df_prateleira.columns.str.strip()

            print(f"Arquivo {arquivo} lido com sucesso")
        return df_prateleira
    except PermissionError:
        messagebox.showerror("Erro", "Erro ao ler os arquivos. Verifique se o arquivo Gabarito Prateleira não esteja aberto")
        raise Exception('Erro ao ler arquivo Gabarito Prateleira')

def encontrar_caminho_orgao_sigla():
    caminho_pasta = var_caminho_pasta.get()
    nome_orgao_sigla = [f for f in os.listdir(caminho_pasta) if (f.startswith('Gabarito Órgão-Sigla') 
                            or f.startswith('Gabarito Órgão Sigla')
                            or f.startswith('Gabarito órgão-sigla')
                            or f.startswith('Gabarito Orgão-Sigla'))
                            and f.endswith('xlsx')]

    if not nome_orgao_sigla:
        print("Nenhum arquivo de 'Gabarito Órgão-Sigla' encontrado.")
        return None

    try:
        for arquivo in nome_orgao_sigla:
            caminho_orgao_sigla = os.path.join(caminho_pasta, arquivo)
            df_orgao_sigla = pd.read_excel(caminho_orgao_sigla)
            df_orgao_sigla.columns = df_orgao_sigla.columns.str.strip()

            print(f"Arquivo {arquivo} lido com sucesso")
        return df_orgao_sigla
    except PermissionError:
        messagebox.showerror("Erro", "Erro ao ler os arquivos. Verifique se o arquivo Gabarito Órgão-Sigla esteja aberto")
        raise Exception('Erro ao ler Gabarito Órgão-Sigla')

# A partir daqui o código é referente à janela 

def selecionar_arquivo():
    caminho_pasta = askdirectory(title='Selecione a pasta com os arquivos')
    var_caminho_pasta.set(caminho_pasta)    
    if caminho_pasta:
        label_pasta_selecionada['text'] = f"* Pasta selecionada: {os.path.basename(caminho_pasta)}"

janela = tk.Tk()
janela.geometry('400x240')
janela.title('Valor Faturado')
janela.resizable(False, False)
texto = tk.Label(janela, text='Selecione a pasta com os arquivos:')
texto.grid(column=1, row=0, padx=10, pady=10, sticky='w')
janela.grid_columnconfigure(1, weight=2)

var_caminho_pasta = tk.StringVar()
botao_selecionararquivo = tk.Button(janela, text="Selecionar", command=selecionar_arquivo)
botao_selecionararquivo.grid(row=1, column=1, padx=10, pady=10, sticky='nsew')
label_pasta_selecionada = tk.Label(janela, text='* Nenhuma pasta selecionada', fg='blue')
label_pasta_selecionada.grid(row=2, column=1, sticky='nsew')

botao_processar = tk.Button(janela, text='Processar', command=iniciar_processamento)
botao_processar.grid(column=1, row=3, columnspan=1, padx=10, pady=10, ipady=5, sticky='nsew')

# Criar o label e a barra de progresso, mas não posicioná-los ainda
loading_label = tk.Label(janela, text="", fg='green')
progress_bar = ttk.Progressbar(janela, orient='horizontal', mode='indeterminate')

# Tratando as imagens que farão parte do botão:
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Ícone da janela
image_path = resource_path("avmg.png")
imagem = Image.open(image_path)
imagem = imagem.resize((32, 32), Image.LANCZOS)
icone = ImageTk.PhotoImage(imagem)
janela.iconphoto(True, icone)

# Botão de ajuda
def ajuda():
    messagebox.showinfo("Informações importantes",
                        "Selecione a pasta que contenha os arquivos:\n\nApuração do Faturamento\nGabarito Prateleira - SIAD\nGabarito Órgão - Sigla")

icon_path = resource_path("botao_de_ajuda_transparente.png")
help_icon = Image.open(icon_path)
help_icon = help_icon.resize((22, 22), Image.LANCZOS)
help_icon = ImageTk.PhotoImage(help_icon)
help_button = tk.Button(janela, image=help_icon, command=ajuda, borderwidth=0, bg='#f0f0f0', activebackground='#f0f0f0')
help_button.grid(row=6, column=5,padx=1, sticky='e')

janela.mainloop()
